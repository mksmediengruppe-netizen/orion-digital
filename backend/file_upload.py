"""
ORION File Upload Module
=========================
Handles file uploads via drag & drop in chat.
Supports: PDF, Excel, images, code files, archives.
Extracts text content for LLM context injection.
"""

import os
import uuid
import json
import logging
import mimetypes
import hashlib
from typing import Dict, List, Optional, Tuple
from datetime import datetime, timezone
from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename

logger = logging.getLogger("file_upload")

UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "/var/www/orion/backend/data/uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
ALLOWED_EXTENSIONS = {
    # Documents
    "pdf", "doc", "docx", "txt", "md", "rtf", "odt",
    # Spreadsheets
    "xlsx", "xls", "csv", "tsv",
    # Images
    "png", "jpg", "jpeg", "gif", "webp", "svg", "bmp",
    # Code
    "py", "js", "ts", "jsx", "tsx", "html", "css", "json", "xml", "yaml", "yml",
    "java", "c", "cpp", "h", "go", "rs", "rb", "php", "sql", "sh", "bash",
    # Archives
    "zip", "tar", "gz", "7z",
    # Data
    "jsonl", "parquet",
}


class FileUploadManager:
    """Manages file uploads and content extraction."""

    def __init__(self, upload_dir: str = None):
        self.upload_dir = upload_dir or UPLOAD_DIR
        os.makedirs(self.upload_dir, exist_ok=True)

    def save_file(self, file_storage, chat_id: str = None, user_id: str = None) -> Dict:
        """Save an uploaded file and extract metadata."""
        filename = secure_filename(file_storage.filename or "unnamed")
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if ext not in ALLOWED_EXTENSIONS:
            return {"success": False, "error": f"File type .{ext} not allowed"}

        # Read file data
        file_data = file_storage.read()
        if len(file_data) > MAX_FILE_SIZE:
            return {"success": False, "error": f"File too large (max {MAX_FILE_SIZE // 1024 // 1024}MB)"}

        # Generate unique filename
        file_hash = hashlib.md5(file_data).hexdigest()[:8]
        unique_name = f"{uuid.uuid4().hex[:8]}_{file_hash}_{filename}"

        # Save to disk
        if chat_id:
            chat_dir = os.path.join(self.upload_dir, chat_id)
            os.makedirs(chat_dir, exist_ok=True)
            filepath = os.path.join(chat_dir, unique_name)
        else:
            filepath = os.path.join(self.upload_dir, unique_name)

        with open(filepath, "wb") as f:
            f.write(file_data)

        # Extract text content
        text_content = self._extract_text(filepath, ext, file_data)

        mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"

        return {
            "success": True,
            "file": {
                "id": unique_name.split("_")[0],
                "filename": filename,
                "stored_name": unique_name,
                "path": filepath,
                "size": len(file_data),
                "mime_type": mime_type,
                "extension": ext,
                "chat_id": chat_id,
                "user_id": user_id,
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "text_content": text_content[:10000] if text_content else None,
                "text_length": len(text_content) if text_content else 0,
                "url": f"/api/uploads/{chat_id}/{unique_name}" if chat_id else f"/api/uploads/{unique_name}",
            },
        }

    def _extract_text(self, filepath: str, ext: str, data: bytes = None) -> Optional[str]:
        """Extract text content from uploaded file."""
        try:
            # Plain text files
            if ext in ("txt", "md", "csv", "tsv", "json", "xml", "yaml", "yml",
                       "py", "js", "ts", "jsx", "tsx", "html", "css", "sql", "sh",
                       "bash", "java", "c", "cpp", "h", "go", "rs", "rb", "php",
                       "jsonl"):
                return data.decode("utf-8", errors="replace") if data else open(filepath).read()

            # PDF
            if ext == "pdf":
                return self._extract_pdf(filepath)

            # Excel
            if ext in ("xlsx", "xls"):
                return self._extract_excel(filepath)

            # DOCX
            if ext == "docx":
                return self._extract_docx(filepath)

        except Exception as e:
            logger.debug(f"Text extraction failed for {ext}: {e}")

        return None

    def _extract_pdf(self, filepath: str) -> Optional[str]:
        """Extract text from PDF."""
        try:
            import subprocess
            result = subprocess.run(
                ["pdftotext", filepath, "-"],
                capture_output=True, text=True, timeout=30
            )
            if result.returncode == 0:
                return result.stdout
        except:
            pass

        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            text = []
            for page in reader.pages[:50]:  # Max 50 pages
                text.append(page.extract_text() or "")
            return "\n".join(text)
        except:
            pass

        return None

    def _extract_excel(self, filepath: str) -> Optional[str]:
        """Extract text from Excel."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            text = []
            for sheet_name in wb.sheetnames[:5]:  # Max 5 sheets
                ws = wb[sheet_name]
                text.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(max_row=500, values_only=True):
                    vals = [str(v) if v is not None else "" for v in row]
                    text.append("\t".join(vals))
            wb.close()
            return "\n".join(text)
        except:
            pass
        return None

    def _extract_docx(self, filepath: str) -> Optional[str]:
        """Extract text from DOCX."""
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            with zipfile.ZipFile(filepath) as z:
                with z.open("word/document.xml") as f:
                    tree = ET.parse(f)
                    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                    paragraphs = tree.findall(".//w:p", ns)
                    text = []
                    for p in paragraphs:
                        runs = p.findall(".//w:t", ns)
                        text.append("".join(r.text or "" for r in runs))
                    return "\n".join(text)
        except:
            pass
        return None

    def get_file(self, filename: str, chat_id: str = None) -> Optional[str]:
        """Get file path by name."""
        if chat_id:
            path = os.path.join(self.upload_dir, chat_id, filename)
        else:
            path = os.path.join(self.upload_dir, filename)
        return path if os.path.exists(path) else None

    def list_files(self, chat_id: str = None) -> List[Dict]:
        """List uploaded files."""
        target_dir = os.path.join(self.upload_dir, chat_id) if chat_id else self.upload_dir
        if not os.path.exists(target_dir):
            return []

        files = []
        for f in sorted(os.listdir(target_dir), reverse=True):
            path = os.path.join(target_dir, f)
            if os.path.isfile(path):
                files.append({
                    "filename": f,
                    "size": os.path.getsize(path),
                    "uploaded_at": datetime.fromtimestamp(os.path.getctime(path)).isoformat(),
                    "url": f"/api/uploads/{chat_id}/{f}" if chat_id else f"/api/uploads/{f}",
                })
        return files

    def delete_file(self, filename: str, chat_id: str = None) -> bool:
        path = self.get_file(filename, chat_id)
        if path:
            os.remove(path)
            return True
        return False


# ── Singleton ──
_manager: Optional[FileUploadManager] = None

def get_upload_manager() -> FileUploadManager:
    global _manager
    if _manager is None:
        _manager = FileUploadManager()
    return _manager


# ── Flask Routes ──

def register_upload_routes(app: Flask):
    """Register file upload API routes."""

    @app.route("/api/upload", methods=["POST"])
    def phase6_upload_file():
        if "file" not in request.files:
            return jsonify({"success": False, "error": "No file provided"}), 400

        file = request.files["file"]
        chat_id = request.form.get("chat_id")
        user_id = getattr(request, "user_id", None)

        mgr = get_upload_manager()
        result = mgr.save_file(file, chat_id=chat_id, user_id=user_id)
        return jsonify(result), 200 if result.get("success") else 400

    @app.route("/api/upload/multiple", methods=["POST"])
    def phase6_upload_multiple():
        files = request.files.getlist("files")
        if not files:
            return jsonify({"success": False, "error": "No files provided"}), 400

        chat_id = request.form.get("chat_id")
        user_id = getattr(request, "user_id", None)
        mgr = get_upload_manager()

        results = []
        for f in files[:10]:  # Max 10 files at once
            result = mgr.save_file(f, chat_id=chat_id, user_id=user_id)
            results.append(result)

        return jsonify({
            "success": True,
            "files": [r.get("file") for r in results if r.get("success")],
            "errors": [r.get("error") for r in results if not r.get("success")],
        })

    @app.route("/api/uploads/<path:filepath>")
    def phase6_serve_upload(filepath):
        full_path = os.path.join(UPLOAD_DIR, filepath)
        if os.path.exists(full_path) and os.path.isfile(full_path):
            return send_file(full_path)
        return jsonify({"error": "Not found"}), 404

    @app.route("/api/uploads", methods=["GET"])
    def phase6_list_uploads():
        chat_id = request.args.get("chat_id")
        mgr = get_upload_manager()
        files = mgr.list_files(chat_id=chat_id)
        return jsonify({"files": files, "count": len(files)})

    @app.route("/api/uploads/<filename>", methods=["DELETE"])
    def phase6_delete_upload(filename):
        chat_id = request.args.get("chat_id")
        mgr = get_upload_manager()
        if mgr.delete_file(filename, chat_id):
            return jsonify({"deleted": True})
        return jsonify({"error": "Not found"}), 404

    logger.info("[UPLOAD] Routes registered")
